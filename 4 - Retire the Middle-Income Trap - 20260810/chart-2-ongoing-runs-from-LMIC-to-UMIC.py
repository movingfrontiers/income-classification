# ---- shared data preparation: builds oghist.json, gni.json, runs_base3.json if missing ----
import os, json, csv as _csv, re as _re
import openpyxl as _oxl
OGHIST_XLSX = '/mnt/user-data/uploads/OGHIST_2026_07_01.xlsx'
WDI_CSV = '/mnt/user-data/uploads/0522fd40-d490-4766-b083-81e9c0ae2068_Data.csv'
BASE_XLSX = '/mnt/user-data/uploads/OHIST_base.xlsx'
if not os.path.exists('oghist.json'):
    _wb = _oxl.load_workbook(OGHIST_XLSX)
    _ws = _wb['Country Analytical History']
    _yc = {}
    for _j in range(1, 260):
        _v = _ws.cell(6, _j).value
        if isinstance(_v, (int, float)) and 1987 <= int(_v) <= 2025:
            _yc[_j] = int(_v)
    _data, _names = {}, {}
    for _i in range(7, _ws.max_row + 1):
        _code, _name = _ws.cell(_i, 1).value, _ws.cell(_i, 2).value
        if not _code or not _name: continue
        _code = str(_code).strip()
        _data[_code] = {str(_y): (str(_ws.cell(_i, _j).value).strip() if _ws.cell(_i, _j).value is not None else '..')
                        for _j, _y in _yc.items()}
        _names[_code] = str(_name).strip()
    _tws = _wb['Thresholds']
    _ty = {_j: int(_tws.cell(7, _j).value) for _j in range(2, 56)
           if isinstance(_tws.cell(7, _j).value, (int, float)) and 1987 <= int(_tws.cell(7, _j).value) <= 2025}
    def _num(_s, _k=0):
        _m = _re.findall(r'[\d,]+', str(_s))
        return float(_m[_k].replace(',', '')) if len(_m) > _k else None
    _thr = {str(_y): {'LM': _num(_tws.cell(22, _j).value), 'UM': _num(_tws.cell(23, _j).value),
                      'H': _num(_tws.cell(24, _j).value)} for _j, _y in _ty.items()}
    json.dump({'years': sorted(set(_yc.values())), 'data': _data, 'names': _names, 'thr': _thr},
              open('oghist.json', 'w'))
if not os.path.exists('gni.json'):
    _g = {}
    with open(WDI_CSV, newline='', encoding='utf-8-sig') as _f:
        _rd = _csv.DictReader(_f)
        _ycols = [c for c in _rd.fieldnames if _re.match(r'^\d{4} \[YR\d{4}\]$', c or '')]
        for _row in _rd:
            if _row.get('Series Code') != 'NY.GNP.PCAP.CD': continue
            _d = {str(int(c[:4])): float(_row[c]) for c in _ycols if _row[c] not in ('..', '', None)}
            if _d: _g[_row['Country Code']] = _d
    json.dump(_g, open('gni.json', 'w'))
if not os.path.exists('runs_base3.json'):
    _wb2 = _oxl.load_workbook(BASE_XLSX)
    _ws2 = _wb2['Classifications']
    _hdr = [c.value for c in _ws2[1]]
    _y0 = _hdr.index('1987') + 1
    _years = list(range(1987, 2026))
    _runs = {'leg1': [], 'leg2': []}
    for _i in range(2, _ws2.max_row + 1):
        _name, _code = _ws2.cell(_i, 1).value, _ws2.cell(_i, 2).value
        if not _name: continue
        if _ws2.cell(_i, _hdr.index('Disqualified') + 1).value or _ws2.cell(_i, _hdr.index('Unselect') + 1).value:
            continue
        _cls, _span = {}, []
        for _k, _y in enumerate(_years):
            _c = _ws2.cell(_i, _y0 + _k)
            _cls[_y] = str(_c.value).strip() if _c.value is not None else '..'
            _fl = _c.fill
            if _fl and _fl.fill_type == 'solid' and str(_fl.fgColor.rgb) in ('FF92D050', 'FFFFFF00'):
                _span.append(_y)
        if not _span: continue
        _s, _e = _span[0], _span[-1]
        _first = 'LM' if _cls[_s] in ('LM', 'LM*') else _cls[_s]
        _leg, _tgt = ('leg1', 'UM') if _first == 'LM' else ('leg2', 'H')
        _endc = _cls[_e]
        _comp = (_endc == _tgt) and (_e < 2025 or _cls[_e - 1] == _tgt)
        _runs[_leg].append(dict(code=_code, name=_name, start=_s, end=_e,
                                status='completed' if _comp else 'incomplete'))
    _gni = {c: {int(y): v for y, v in d.items()} for c, d in json.load(open('gni.json')).items()}
    _thr2 = {int(k): v for k, v in json.load(open('oghist.json'))['thr'].items()}
    for _leg, _tk in [('leg1', 'UM'), ('leg2', 'H')]:
        for _r in _runs[_leg]:
            if _r['status'] != 'incomplete': continue
            _gd = _gni.get(_r['code'], {})
            _av = [y for y in _gd if _r['start'] <= y <= 2025]
            if _av:
                _yl = max(_av)
                _r['ratio'] = 100 * _gd[_yl] / _thr2[_yl][_tk]
    json.dump(_runs, open('runs_base3.json', 'w'))
    json.dump(_runs, open('runs_base2.json', 'w'))
# ---- end shared data preparation ----

# Core builder for the LMIC->UMIC / UMIC->HIC facet charts (final specification).
# Requires oghist.json, gni.json (from _preamble) and OHIST_base.xlsx (population column).
import json, math
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import NullLocator, NullFormatter
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from PIL import Image, ImageDraw, ImageFont
import openpyxl

plt.rcParams.update({
    'font.family': 'DejaVu Sans', 'font.size': 18,
    'xtick.labelsize': 18.9, 'ytick.labelsize': 18.9,
    'axes.edgecolor': '#999', 'axes.linewidth': 1.0,
    'axes.facecolor': 'white', 'figure.facecolor': 'white',
})
C_DONE = '#1F4E79'
ACC = '#B2503B'
CAP_FS = 12.9
WM_FS = CAP_FS * 1.2
WM_TXT = 'movingfrontiers.substack.com'
DPI = 200
FONT_DIR = '/usr/share/fonts/truetype/dejavu/'

_og = json.load(open('oghist.json'))
gni = {c: {int(y): v for y, v in d.items()} for c, d in json.load(open('gni.json')).items()}
thr = {int(k): v for k, v in _og['thr'].items()}
DISS = {'CSK', 'SUN', 'YUG', 'YUGf', 'ANT'}

def _norm(v): return 'LM' if v == 'LM*' else v

def make_runs():
    out = {'leg1': [], 'leg2': []}
    for code, cls_raw in _og['data'].items():
        if code in DISS or code == 'GNQ':
            continue
        cls = {int(y): _norm(v) for y, v in cls_raw.items()}
        name = _og['names'][code]
        for leg, bracket, below, target in [('leg1', 'LM', {'L'}, {'UM', 'H'}),
                                            ('leg2', 'UM', {'L', 'LM'}, {'H'})]:
            start = None
            for y in range(1988, 2026):
                if cls.get(y) == bracket and cls.get(y - 1) in below:
                    start = y
                    break
            if start is None:
                continue
            end = None
            for y in range(start + 1, 2026):
                if cls.get(y) in target:
                    end = y
                    break
            if end is not None:
                if end - start < 2:
                    continue
                out[leg].append(dict(code=code, name=name, start=start, end=end, status='completed'))
            else:
                if 2025 - start < 2:
                    continue
                out[leg].append(dict(code=code, name=name, start=start, end=2025, status='incomplete'))
    return out

runs = make_runs()

_wbp = openpyxl.load_workbook('/mnt/user-data/uploads/OHIST_base.xlsx')
_wsp = _wbp['Classifications']
_hdrp = [c.value for c in _wsp[1]]
POPM = {}
for _i in range(2, _wsp.max_row + 1):
    _code = _wsp.cell(_i, 2).value
    _p = _wsp.cell(_i, _hdrp.index('Pop. (m)') + 1).value
    if _code and _p is not None and str(_code).strip() not in POPM:
        POPM[str(_code).strip()] = float(_p)
POP_MIN = 10.0

SHORT = {'Bosnia and Herzegovina': 'Bosnia & Herz.', 'Russian Federation': 'Russia',
         'Iran, Islamic Rep.': 'Iran', 'Slovak Republic': 'Slovakia',
         'Dominican Republic': 'Dom. Rep.', 'Papua New Guinea': 'PNG',
         'Egypt, Arab Rep.': 'Egypt', 'Venezuela, RB': 'Venezuela', 'Yemen, Rep.': 'Yemen'}
def sn(n): return SHORT.get(n, n)

def run_ratio(r, tkey):
    g = gni.get(r['code'], {})
    xs, ys = [], []
    for y in range(r['start'], r['end'] + 1):
        v = g.get(y)
        t = thr.get(y, {}).get(tkey)
        xs.append(y - r['start'])
        ys.append(min(100 * v / t, 100.0) if (v and t) else np.nan)
    return np.array(xs, float), np.array(ys, float)

def comp_set(leg):
    tkey = 'UM' if leg == 'leg1' else 'H'
    comp = sorted([r for r in runs[leg] if r['status'] == 'completed'
                   and POPM.get(r['code'], 0) >= POP_MIN],
                  key=lambda r: (r['end'] - r['start'], r['name']))
    return [r for r in comp if not np.all(np.isnan(run_ratio(r, tkey)[1]))], tkey

def ongo_set(leg):
    tkey = 'UM' if leg == 'leg1' else 'H'
    out = []
    for r in runs[leg]:
        if r['status'] != 'incomplete' or POPM.get(r['code'], 0) < POP_MIN:
            continue
        if np.all(np.isnan(run_ratio(r, tkey)[1])):
            continue
        g = gni.get(r['code'], {})
        avail = [y for y in g if r['start'] <= y <= 2025]
        ratio = 100 * g[max(avail)] / thr[max(avail)][tkey] if avail else 0
        out.append(dict(r, ratio=ratio))
    def last_obs(r):
        g = gni.get(r['code'], {})
        avail = [y for y in g if r['start'] <= y <= 2025]
        return max(avail) if avail else 0
    dropped = sorted(sn(r['name']) for r in out if 2025 - r['start'] <= 3)
    stale = sorted(sn(r['name']) for r in out if 2025 - r['start'] > 3 and last_obs(r) < 2020)
    out = [r for r in out if 2025 - r['start'] > 3 and last_obs(r) >= 2020]
    return out, dropped, stale

YLOW_LEG = {}
for leg in ('leg1', 'leg2'):
    glo = 100
    comp, tkey = comp_set(leg)
    for r in comp:
        glo = min(glo, np.nanmin(run_ratio(r, tkey)[1]))
    for r in ongo_set(leg)[0]:
        ys = run_ratio(r, tkey)[1]
        if not np.all(np.isnan(ys)):
            glo = min(glo, np.nanmin(ys))
    YLOW_LEG[leg] = max(5, math.floor(glo / 5) * 5)
XMAX = 33

NOTE = ("Source: World Bank, historical income classifications (OGHIST) and FY27 analytical classification, July 2026; World "
        "Development Indicators, GNI per capita, Atlas method (current US$), update of 13 July 2026. Note: LMIC = "
        "lower-middle-income country, UMIC = upper-middle-income country, HIC = high-income country. Economies with populations "
        "under 10 million are not shown. Completed runs, one panel per economy, ordered fastest to slowest by duration, and the "
        "fastest, slowest and median runs tagged, and the duration and years of each run shown inside the panel. Episode rules: "
        "a run starts in the first year the economy crossed into the bracket from the category below, and completes in the first "
        "year it reached the target category afterwards; runs shorter than two years are excluded; economies that entered the "
        "classification already in the bracket and where the crossing cannot be observed in 1987 or at a later date are "
        "excluded; Equatorial Guinea is excluded due to extreme volatility. Each panel tracks the economy's Atlas GNI per capita "
        "as a share of the target threshold prevailing in each calendar year, censored at the target line (100). Dips inside a "
        "run reflect temporary reclassifications or income declines along the way. The gray band, identical in every panel, "
        "spans the fastest and slowest completed runs; the solid light gray line inside the band is the median completed run.")

NOTE_ONGO = ("Source: World Bank, historical income classifications (OGHIST) and FY27 analytical classification, July 2026; World "
        "Development Indicators, GNI per capita, Atlas method (current US$), update of 13 July 2026. Note: LMIC = "
        "lower-middle-income country, UMIC = upper-middle-income country, HIC = high-income country. Economies with populations "
        "under 10 million are not shown. {DROPPED}Ongoing runs, one panel per economy, ordered by how far each economy runs "
        "behind the median completed run: at its latest income level, the horizontal distance between its years in the bracket "
        "and the years the median run needed to reach that level, from smallest to largest, with the duration and years of each "
        "run shown inside the panel. Episode rules: a run starts in the first year the economy crossed into the bracket from the "
        "category below, and would complete in the first year it reached the target category afterwards; runs shorter than two "
        "years are excluded; economies that entered the classification already in the bracket and where the crossing cannot be "
        "observed in 1987 or at a later date are excluded, as are economies without usable GNI data; Equatorial Guinea is "
        "excluded due to extreme volatility. Each panel tracks the economy's Atlas GNI per capita as a share of the target "
        "threshold prevailing in each calendar year, censored at the target line (100); paths can run along the line where "
        "income stands at or above the threshold without the reclassification having occurred, as where the World Bank exercised "
        "discretion citing Atlas-method limitations under high inflation (T\u00fcrkiye, FY27). Paths end at the latest available "
        "observation. The gray band, identical in every panel, spans the fastest and slowest completed runs of this leg; the "
        "solid light gray line inside the band is the median completed run.")

CLOSERS = ["", " Axes match across panels.", " Panels share identical axes.",
           " Axes are identical across panels.", " All panels share identical axes.",
           " Axes are identical in all panels shown.", " All panels share the same axes throughout.",
           " Axes are identical across all panels of this chart.",
           " All panels share identical horizontal and vertical axes.",
           " All panels share identical horizontal and vertical axes throughout.",
           " All panels share identical horizontal and vertical axes, drawn to the same scale.",
           " All panels of this chart share identical horizontal and vertical axes, drawn to the same scale.",
           " All panels of this chart and its companion share identical horizontal and vertical axes, drawn to the same scale."]

def facet_chart(leg, title, subtitle, fname, ongoing=False):
    comp, tkey = comp_set(leg)
    dropped, stale = [], []
    if ongoing:
        panel_runs, dropped, stale = ongo_set(leg)
    else:
        panel_runs = comp
    n = len(panel_runs)
    ncol = 4 if ongoing else 3
    nrow = math.ceil(n / ncol)
    height = 3.1 * nrow + (6.95 if ongoing else 5.8)
    bot_in = 6.45 if ongoing else 5.35
    fig, axes = plt.subplots(nrow, ncol, figsize=(12.6, height), dpi=DPI, sharex=True, sharey=True)
    fig.subplots_adjust(top=1 - 0.35 / height, bottom=bot_in / height, left=0.095, right=0.972,
                        hspace=0.40, wspace=0.13)
    fast, slow = comp[0], comp[-1]
    med = comp[(len(comp) - 1) // 2]
    fX, fY = run_ratio(fast, tkey)
    sX, sY = run_ratio(slow, tkey)
    mX, mY = run_ratio(med, tkey)
    mok = ~np.isnan(mY)
    mXo, mYo = mX[mok], mY[mok]
    def x_med_at(yv):
        if yv <= mYo[0]:
            return 0.0
        for a in range(1, len(mYo)):
            if mYo[a] >= yv and mYo[a - 1] < yv:
                return float(mXo[a - 1] + (yv - mYo[a - 1]) / (mYo[a] - mYo[a - 1]) * (mXo[a] - mXo[a - 1]))
        return float(mXo[-1])
    if ongoing:
        for r in panel_runs:
            xs_, ys_ = run_ratio(r, tkey)
            idx_ = np.where(~np.isnan(ys_))[0]
            xl, yl = float(xs_[idx_[-1]]), float(ys_[idx_[-1]])
            r['dist'] = xl - x_med_at(yl)
        panel_runs.sort(key=lambda r: (r['dist'], r['name']))
    fok, sok = ~np.isnan(fY), ~np.isnan(sY)
    Xc = list(range(0, int(sX[-1]) + 1))
    UP = [min(float(np.interp(x, fX[fok], fY[fok])), 100) if x <= fX[fok][-1] else 100 for x in Xc]
    LO = [min(float(np.interp(x, sX[sok], sY[sok])), u) for x, u in zip(Xc, UP)]
    for k, ax in enumerate(axes.flat):
        if k >= n:
            ax.axis('off'); continue
        r = panel_runs[k]
        dur = (2025 - r['start']) if ongoing else (r['end'] - r['start'])
        ax.fill_between(Xc, LO, UP, color='#999', alpha=0.15, lw=0, zorder=1)
        ax.plot(mX, mY, color='#bbb', lw=1.1, zorder=1.6)
        ax.axhline(100, color='#666', lw=1.1, ls=(0, (4, 3)), zorder=2)
        xs, ys = run_ratio(r, tkey)
        line_c = ACC if ongoing else C_DONE
        ax.plot(xs, ys, color=line_c, lw=2.6, zorder=3)
        idx = np.where(~np.isnan(ys))[0]
        if ongoing:
            ax.plot(xs[idx[-1]], ys[idx[-1]], 'o', mfc='white', mec=line_c, mew=1.6, ms=6.0, zorder=4)
        else:
            ax.plot(xs[idx[0]], ys[idx[0]], 'o', mfc='white', mec=line_c, mew=1.6, ms=5.2, zorder=4)
            ax.plot(xs[idx[-1]], ys[idx[-1]], 'o', color=line_c, ms=5.6, zorder=4)
        YLOW = YLOW_LEG[leg]
        ax.set_xlim(-0.8, XMAX)
        ax.set_ylim(YLOW, 104)
        ax.set_xticks([0, 10, 20, 30])
        ax.set_yticks([v for v in (25, 50, 75, 100) if v > YLOW])
        ax.yaxis.set_minor_locator(NullLocator()); ax.yaxis.set_minor_formatter(NullFormatter())
        ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
        tag, col = '', '#333'
        if not ongoing:
            if r is fast: tag, col = ' \u00b7 fastest', ACC
            elif r is slow: tag, col = ' \u00b7 slowest', ACC
            elif r is med: tag, col = ' \u00b7 median', ACC
        ax.set_title(f"{sn(r['name'])}{tag}", fontsize=18.9, fontweight='bold',
                     color=col, loc='left', pad=6)
        rt = ax.text(0.96, 0.05, f"{r['start']}\u2013{r['end']}", transform=ax.transAxes,
                     fontsize=12.9, color='#555', ha='right', va='bottom', zorder=6)
        ax.annotate(f"{dur} yr" + ('s' if dur != 1 else ''), xy=(0, 1), xycoords=rt,
                    xytext=(0, 3), textcoords='offset points', fontsize=12.9, color='#555',
                    ha='left', va='bottom', zorder=6)
    for j in range(ncol):
        k = (nrow - 1) * ncol + j
        target_ax = axes[-1][j] if k < n else axes[-2][j]
        target_ax.set_xlabel('Years since entry', fontsize=18.9)
        target_ax.tick_params(labelbottom=True)
    fig.text(0.016, (fig.subplotpars.top + fig.subplotpars.bottom) / 2,
             'GNI per capita, % of contemporaneous target line (censored at 100)',
             fontsize=18.9, rotation=90, va='center', ha='center')
    if ongoing:
        h_run = Line2D([], [], color=ACC, lw=2.6, marker='o', mfc='white', mew=1.6, ms=6.0,
                       label='Ongoing run (open dot at latest data)')
    else:
        h_run = Line2D([], [], color=C_DONE, lw=2.6, marker='o', ms=5.6,
                       label='Completed run (open dot at entry, solid dot at completion)')
    handles = [h_run,
               Patch(facecolor='#999', alpha=0.2, edgecolor='none',
                     label=f'Range of completed runs: fastest ({sn(fast["name"])}) to slowest ({sn(slow["name"])})'),
               Line2D([], [], color='#bbb', lw=1.3,
                      label=f'Median completed run ({sn(med["name"])})')]
    fig.canvas.draw()
    ren = fig.canvas.get_renderer()
    W, H = fig.get_size_inches() * fig.dpi
    xl_low = min(a.xaxis.label.get_window_extent(ren).y0 for row in axes for a in row
                 if a.get_visible() and a.xaxis.label.get_text())
    extra = (0.5 * 18.9 * 1.35 / 72 * fig.dpi / H) if ongoing else 0.0
    lgd = fig.legend(handles=handles, loc='upper center',
                     bbox_to_anchor=(0.5, xl_low / H - 0.006 - extra),
                     ncol=1, frameon=False, fontsize=18.9, handletextpad=0.5, columnspacing=1.1)
    fig.canvas.draw()
    lg_bb = lgd.get_window_extent(ren)

    probe = fig.text(0.5, 0.5, '0', fontsize=CAP_FS)
    fig.canvas.draw()
    char_w = probe.get_window_extent(ren).width
    def width_of(s, fs=CAP_FS):
        probe.set_text(s); probe.set_fontsize(fs)
        return probe.get_window_extent(ren).width
    line_h = CAP_FS * 1.35 / 72 * fig.dpi
    x0_px = 0.012 * W
    right_px = W - 70
    full_w = right_px - x0_px
    wm_w = width_of(WM_TXT, WM_FS)
    gap_target = width_of('movi', WM_FS)
    lim_last = (right_px - wm_w - x0_px) - gap_target
    SAFE = 6
    wcache = {}
    def wof(w):
        if w not in wcache: wcache[w] = width_of(w)
        return wcache[w]
    w_space = width_of('0 0') - 2 * width_of('0')
    def lw_model(ws):
        return sum(wof(w) for w in ws) + (len(ws) - 1) * w_space
    def greedy_m(ws_list, maxw):
        out, cur = [], []
        for w in ws_list:
            if cur and lw_model(cur + [w]) > maxw:
                out.append(cur); cur = [w]
            else:
                cur.append(w)
        out.append(cur)
        return out
    if ongoing:
        if len(dropped) == 1:
            dsent = dropped[0] + ' is not shown, having entered the bracket three or fewer years ago. '
        elif dropped:
            dsent = ', '.join(dropped[:-1]) + ' and ' + dropped[-1] + ' are not shown, having entered the bracket three or fewer years ago. '
        else:
            dsent = ''
        if len(stale) == 1:
            dsent += stale[0] + ' is not shown, as its latest income observation is more than five years old. '
        elif stale:
            dsent += ', '.join(stale[:-1]) + ' and ' + stale[-1] + ' are not shown, as their latest income observations are more than five years old. '
        base_note = NOTE_ONGO.replace('{DROPPED}', dsent)
    else:
        base_note = NOTE
    marker = ' Note: '
    p_src, p_note = base_note.split(marker, 1)
    src_lines = [' '.join(ws) for ws in greedy_m(p_src.split(), full_w - SAFE)]
    chosen = None
    for closer in CLOSERS:
        words = ('Note: ' + p_note + closer).split()
        head_all = greedy_m(words, full_w - SAFE)
        for k in (1, 2):
            if len(head_all) < 2 + k:
                continue
            tail_words = [w for ln in head_all[-k:] for w in ln]
            tl = greedy_m(tail_words, lim_last - SAFE)
            if len(tl) == 2 and lw_model(tl[0]) >= 0.70 * lim_last:
                chosen = head_all[:-k] + tl
                break
        if chosen:
            break
    assert chosen is not None, f'{fname}: no closer yields the last-two-lines layout'
    lines = src_lines + [' '.join(ws) for ws in chosen]
    for ln in lines:
        assert width_of(ln) <= full_w + 1, f'{fname}: line overflow'
    cap_text = '\n'.join(lines)
    cap_top_px = lg_bb.y0 - 0.80 * line_h
    cap = fig.text(0.012, cap_top_px / H, cap_text, fontsize=CAP_FS, color='#666',
                   ha='left', va='top', linespacing=1.35)
    fig.canvas.draw()
    cb = cap.get_window_extent(ren)
    bprobe = fig.text(0.5, 0.5, '0', fontsize=CAP_FS, va='baseline')
    fig.canvas.draw()
    desc_cap = 0.5 * H - bprobe.get_window_extent(ren).y0
    bprobe.set_fontsize(WM_FS); fig.canvas.draw()
    desc_wm = 0.5 * H - bprobe.get_window_extent(ren).y0
    bprobe.remove()
    baseline_px = cb.y0 + desc_cap
    wm = fig.text(right_px / W, baseline_px / H, WM_TXT, fontsize=WM_FS, color='#999999',
                  ha='right', va='baseline')
    fig.canvas.draw()
    wb = wm.get_window_extent(ren)
    gap_px = min(wb.x0 - (x0_px + width_of(lines[-1])), wb.x0 - (x0_px + width_of(lines[-2])))
    assert gap_px >= gap_target - 2, f'{fname}: watermark gap {gap_px:.1f}px < movi width ({gap_target:.1f})'
    assert width_of(lines[-2]) >= 0.70 * lim_last, f'{fname}: second-to-last line too short'
    nxt = lines[-2].split()[0]
    assert width_of(lines[-3] + ' ' + nxt) > full_w - 8, f'{fname}: third-to-last line not maximally full'
    wm_base = wb.y0 + desc_wm
    assert abs(wm_base - baseline_px) <= 0.0015 * H + 0.5, f'{fname}: baseline mismatch'
    assert cb.y0 > 0.004 * H, f'{fname}: caption clipped at bottom ({cb.y0:.1f}px)'
    gap_lines = (lg_bb.y0 - cb.y1) / line_h
    assert 0.6 <= gap_lines <= 1.0, f'{fname}: caption top gap {gap_lines:.2f} line-heights'
    probe.remove()
    fig.savefig('/tmp/_stage.png')
    plt.close(fig)

    im = Image.open('/tmp/_stage.png').convert('RGB')
    a = np.array(im)
    ink = (a < 250).any(axis=2)
    rows = np.where(ink.any(axis=1))[0]
    im = im.crop((0, max(0, rows[0] - 10), im.width, min(im.height, rows[-1] + 18)))
    tfont = ImageFont.truetype(FONT_DIR + 'DejaVuSans-Bold.ttf', int(31.5 * DPI / 72))
    sfont = ImageFont.truetype(FONT_DIR + 'DejaVuSans.ttf', int(21.3 * DPI / 72))
    pad_top, gap, pad_bot = 46, 18, 38 + int(21.3 * DPI / 72)
    th = tfont.getbbox(title)[3]
    sh = sfont.getbbox(subtitle)[3]
    x_ttl = int(0.03 * im.width)
    max_w = im.width - 2 * x_ttl
    sub_lines, cur = [], ''
    for w in subtitle.split():
        t = (cur + ' ' + w).strip()
        if sfont.getlength(t) <= max_w:
            cur = t
        else:
            sub_lines.append(cur); cur = w
    sub_lines.append(cur)
    sub_lh = int(sh * 1.30)
    band_h = pad_top + th + gap + sub_lh * len(sub_lines) + pad_bot
    band = Image.new('RGB', (im.width, band_h), 'white')
    d = ImageDraw.Draw(band)
    d.text((x_ttl, pad_top), title, font=tfont, fill='#2d2d2d')
    for si, sl in enumerate(sub_lines):
        d.text((x_ttl, pad_top + th + gap + si * sub_lh), sl, font=sfont, fill='#646464')
    out = Image.new('RGB', (im.width, band_h + im.height), 'white')
    out.paste(band, (0, 0))
    out.paste(im, (0, band_h))
    out.save(fname)
    print(f'{fname}: OK ({n} panels, median {sn(med["name"])}, wm gap {gap_px:.0f}px, cap gap {gap_lines:.2f} lines)')

import csv as _csvmod

def med_inverse(comp, tkey):
    med = comp[(len(comp) - 1) // 2]
    mX, mY = run_ratio(med, tkey)
    mok = ~np.isnan(mY)
    mXo, mYo = mX[mok], mY[mok]
    def x_med_at(yv):
        if yv <= mYo[0]: return 0.0
        for a in range(1, len(mYo)):
            if mYo[a] >= yv and mYo[a - 1] < yv:
                return float(mXo[a - 1] + (yv - mYo[a - 1]) / (mYo[a] - mYo[a - 1]) * (mXo[a] - mXo[a - 1]))
        return float(mXo[-1])
    return med, x_med_at

def write_chart_csv(fname, leg, ongoing):
    tkey = 'UM' if leg == 'leg1' else 'H'
    comp, _ = comp_set(leg)
    med, x_med_at = med_inverse(comp, tkey)
    fast, slow = comp[0], comp[-1]
    if ongoing:
        rows_src, dropped, stale = ongo_set(leg)
        for r in rows_src:
            xs_, ys_ = run_ratio(r, tkey)
            i = np.where(~np.isnan(ys_))[0]
            r['latest_pct'] = float(ys_[i[-1]])
            r['dist'] = float(xs_[i[-1]]) - x_med_at(float(ys_[i[-1]]))
        rows_src.sort(key=lambda r: (r['dist'], r['name']))
    else:
        rows_src = comp
    hdr = ['rank', 'economy', 'iso3', 'status', 'run_start', 'run_end', 'duration_years', 'tag',
           'year', 'years_since_entry', 'gni_per_capita_atlas_usd',
           ('umic_entry' if leg == 'leg1' else 'hic') + '_threshold_usd',
           'income_pct_of_line', 'income_pct_censored_100']
    if ongoing:
        hdr += ['latest_income_pct_of_line', 'distance_to_median_years']
    out = []
    for rank, r in enumerate(rows_src, start=1):
        tag = ''
        if not ongoing:
            if r is fast: tag = 'fastest'
            elif r is slow: tag = 'slowest'
            elif r is med: tag = 'median'
        dur = (2025 - r['start']) if ongoing else (r['end'] - r['start'])
        g = gni.get(r['code'], {})
        for y in range(r['start'], r['end'] + 1):
            v = g.get(y)
            t = thr.get(y, {}).get(tkey)
            raw = round(100 * v / t, 2) if (v and t) else ''
            cen = round(min(100 * v / t, 100.0), 2) if (v and t) else ''
            row = [rank, sn(r['name']), r['code'], 'ongoing' if ongoing else 'completed',
                   r['start'], r['end'], dur, tag, y, y - r['start'],
                   round(v, 1) if v else '', round(t, 1) if t else '', raw, cen]
            if ongoing:
                row += [round(r['latest_pct'], 2), round(r['dist'], 2)]
            out.append(row)
    with open(fname, 'w', newline='', encoding='utf-8-sig') as f:
        w = _csvmod.writer(f)
        w.writerow(hdr)
        w.writerows(out)
    print(fname, 'written:', len(rows_src), 'economies')


facet_chart('leg1', 'Ongoing runs from LMIC to UMIC', 'Income relative to the moving UMIC entry line, economies ordered by distance to the median run', 'chart-2-ongoing-runs-from-LMIC-to-UMIC.png', ongoing=True)
write_chart_csv('chart-2-ongoing-runs-from-LMIC-to-UMIC.csv', 'leg1', True)
